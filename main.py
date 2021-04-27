from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import time
import re

from openpyxl import Workbook

wb = Workbook()




driver = webdriver.Chrome(executable_path=r"C:\Users\Hp\Desktop\Coursera\Driver\chromedriver.exe")

driver.get("https://www.energie-cluster.ch/de/ueber-uns_0/verein-mitgliedschaft/mitgliederliste-2706.html")

html1 = driver.page_source
soup1 = BeautifulSoup(html1, 'html.parser')
lst = soup1.find_all("section", {"class": "article-section"})

for x in range(5):
    sh1 = wb.create_sheet("Sheet", x)
    sh1.cell(row=1, column=1).value = "Company Name"
    sh1.cell(row=1, column=2).value = "Website"
    counter = 0
    r = 2
    for i in lst[x+2].find("p"):
        if not str(i).strip():
            pass
        elif "  "==str(i):
            pass
        elif " " ==str(i):
            pass
        elif "<br/>" == str(i):
            pass
        elif str(i) == "h":
            pass
        elif str(i) == "  ":
            pass
        else:
            counter = counter + 1
            if "www" in str(i):

                print(counter)
                try:
                    e = str(i.get("href"))
                    if "http://webcache.googleusercontent.com/search?q=cache:" in e:
                        e = e.replace("http://webcache.googleusercontent.com/search?q=cache:", "")
                except:
                    e = "None"
                print(e)
                sh1.cell(row=r, column=2).value = e
            elif ".ch" in str(i):
                try:
                    e = str(i.get("href"))
                    if "http://webcache.googleusercontent.com/search?q=cache:" in e:
                        e = e.replace("http://webcache.googleusercontent.com/search?q=cache:", "")
                except:
                    e = "None"
                print(e)
                sh1.cell(row=r, column=2).value = e
            elif "http" in str(i):
                try:
                    e = str(i.get("href"))
                    if "http://webcache.googleusercontent.com/search?q=cache:" in e:
                        e = e.replace("http://webcache.googleusercontent.com/search?q=cache:", "")
                except:
                    e = "None"
                print(e)
                sh1.cell(row=r, column=2).value = e
            else:
                print(str(i))
                sh1.cell(row=r, column=1).value = str(i)
            if counter == 2:
                print("Yes")
                r = r + 1
                counter = 0
        wb.save("C:\\Users\\Hp\\Desktop\\New folder (3)\\test123.xlsx")
